import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def modify_form(file_path):

    # 현재 스크립트가 위치한 디렉토리 경로
    # current_directory = os.path.dirname(os.path.abspath(__file__))

    # 'data' 폴더의 전체 경로
    data_folder_path = os.path.join(file_path)
    print(data_folder_path)

    # 폴더 안의 엑셀 파일 하나씩 불러오기
    file_list = os.listdir(data_folder_path)  # path폴더에 있는 파일을 리스트로 받기
    print(file_list)

    for file_name_raw in file_list:
        file_name = f'{file_path}\\' + file_name_raw
        wb = load_workbook(filename=file_name)  # 엑셀파일 가져오기
        ws = wb.active  # 활성화

        # A열 삭제
        ws.delete_cols(1)

        # 기본 열 너비 설정
        default_width = 13

        # 열 너비 조정
        for col in ws.columns:
            column_letter = get_column_letter(col[0].column)
            max_length = default_width  # 기본 너비 설정

            # 2행 띄어쓰기 포함 문자열 처리: 각 셀의 값을 문자열로 변환하고, 길이를 측정하여 최대 길이를 업데이트
            for cell in col:
                if cell.row == 2:  # 2행만 고려
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))

            # 열 너비 조정 (여유 공간을 위해 2를 추가)
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # 엑셀 파일을 동일한 이름으로 저장
        wb.save(file_name)
        print('엑셀폼 변경 완료')
