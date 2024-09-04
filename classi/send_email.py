import re
from email.message import EmailMessage
from smtplib import SMTP_SSL
from pathlib import Path
from openpyxl import load_workbook
import classi.data_store

class Send:
    def __init__(self, id, pw, df,label_4):
        self.id = id
        self.pw = pw  # 앱 비밀번호 16자리
        self.df = df #email_list.xlsx
        self.label_4 = label_4

    def validate_email(self, email):
        # 이메일 주소 유효성 검사 정규 표현식
        email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(email_regex, email) is not None

    def send_email(self):
        # 엑셀 파일 읽기
        wb = load_workbook(self.df, data_only=True)
        ws = wb.active

        # 이메일 발송
        for row in ws.iter_rows(min_row=2, values_only=True):  # 첫 행은 헤더라고 가정
            recipient = row[0]
            cc_recipient = row[1]
            title = row[2]
            text = row[3]
            attachment = row[4]
            print(recipient)
            print(cc_recipient)
            print(title)
            print(text)
            print(attachment)

            # 데이터가 None일 경우 빈 문자열로 처리
            recipient = recipient if recipient else ""
            cc_recipient = cc_recipient if cc_recipient else ""
            title = title if title else "No Subject"
            text = text if text else "No Body"

            # 이메일 주소를 리스트로 변환
            recipient_list = [email.strip() for email in recipient.split(",")] if isinstance(recipient, str) else []
            cc_recipient_list = [email.strip() for email in cc_recipient.split(",")] if isinstance(cc_recipient,
                                                                                                   str) else []
            # 유효하지 않은 이메일 주소 필터링
            recipient_list = [email for email in recipient_list if self.validate_email(email)]
            cc_recipient_list = [email for email in cc_recipient_list if self.validate_email(email)]

            # 빈 문자열이 있는지 확인
            if not recipient_list or "" in recipient_list:
                self.label_4.setText("받는 메일을 확인해주세요.")
                self.label_4.setStyleSheet("color: red;")
                return

            # 템플릿 생성
            msg = EmailMessage()

            # 보내는 사람 / 받는 사람 / 참조 / 제목 입력
            msg["From"] = self.id
            msg["To"] = ", ".join(recipient_list)
            if cc_recipient_list:
                msg["Cc"] = ", ".join(cc_recipient_list)
            msg["Subject"] = title

            # 본문 구성
            msg.set_content(text)

            # 파일 첨부
            if attachment:
                filenm = Path(attachment).name
                with open(classi.data_store.file_path +'\\'+ attachment, 'rb') as f:
                    msg.add_attachment(f.read(), maintype='application', subtype='octet-stream', filename=filenm)

            # SMTP 서버로 발송
            with SMTP_SSL("smtp.gmail.com", 465) as smtp:
                smtp.login(self.id, self.pw)
                smtp.send_message(msg)

            # 완료 메시지
            self.label_4.setText(f"메일 발송 성공: {', '.join(recipient_list)}")
            self.label_4.setStyleSheet("color: green;")
            print(f"발송 성공: {', '.join(recipient_list)}")